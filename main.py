
import pandas as pd
import numpy as np
import openpyxl as pyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font


input_path=r'./input/input.xlsx'
row_data=pd.read_excel(input_path)
print(row_data.shape)
print(row_data.head())


column_mapping={
    6:(1,'Actua'),
    7:(2,'Act_AVA'),
    8:(3,'Stress'),
    9:(4,'Stress_AVA'),
    10:(5,''),
    11:(6,''),
    12:(7,'Interest Rate'),
    13:(8,'Fx'),
    14:(9,'Equity'),
    15:(10,'Credit Spread'),
    16:(11,'Funds'),
    17:(12,'Commodities')}

df_dict={}
for col_idx, (col_num, col_desc) in column_mapping.items():
    rows = []
    for index,value in enumerate(row_data.iloc[:,col_idx]):
        if pd.isna(value):
            rows.append({
               'Row_num': row_data.iloc[index, 0],
               'Col_num': col_num,
               'Value': np.nan,
               'Row_1': row_data.iloc[index, 1],
               'Counterparty_1': row_data.iloc[index, 2],
               'Counterparty_2': row_data.iloc[index, 3],
               'Counterparty_3': row_data.iloc[index, 4],
               'Credit_Category': row_data.iloc[index, 5],
               'Col_Description': col_desc
                })

    df_dict[f'df_{col_num}'] = pd.DataFrame(rows)
df = pd.concat(df_dict.values(), ignore_index=True)

def template_excel(df,sheet_name,outputpath):
    wbo = pyxl.Workbook()
    sheet=wbo.create_sheet(sheet_name)
    del wbo['Sheet']

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
               cell.style = template.HCOB_STYLES["HL1"]
            else:
               cell.style = template.HCOB_STYLES["BD1"]
               if isinstance(value, (int, float)):
                  cell.number_format = '0'

    sheet.freeze_panes='A2'
    for cell in sheet[2]:
        cell.font = Font(bold=True)
    wbo.save(outputpath)

    return

output=template_excel(df,sheet_name='mapping',outputpath='./CSV-MR-Reserve_mapping.xlsx')
